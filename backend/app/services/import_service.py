import io
import re

from fastapi import HTTPException, status, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.attendee import Attendee
from app.models.attendance import Attendance
from app.models.event import Event
from app.schemas.import_schemas import (
    AttendeeImportPreview,
    AttendeeImportResult,
    ImportRow,
)

# Maximum accepted spreadsheet size (10 MB).
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024

# Maximum rows allowed in a single import.
MAX_ROW_COUNT = 5000

ALLOWED_EXTENSIONS = {".xlsx"}

# Canonical approved fields -> accepted column aliases.
# Matching is case-insensitive and ignores whitespace.
COLUMN_ALIASES = {
    "name": ["name", "full name", "student name"],
    "email": ["email", "email address"],
    "phone_number": ["phone", "phone number", "contact number", "mobile number"],
    "section": ["section"],
    "semester": ["semester"],
}

REQUIRED_COLUMNS = list(COLUMN_ALIASES.keys())

# User-friendly labels for missing-column error messages.
DISPLAY_NAMES = {
    "name": "Name",
    "email": "Email",
    "phone_number": "Phone Number",
    "section": "Section",
    "semester": "Semester",
}

EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _normalize(value: str) -> str:
    """Collapse whitespace and lowercase so header matching is robust."""
    return re.sub(r"\s+", "", value).lower()


# Pre-normalized aliases so lookup compares like-for-like ("emailaddress" matches "email address").
NORMALIZED_ALIASES = {
    canonical: {_normalize(a) for a in aliases}
    for canonical, aliases in COLUMN_ALIASES.items()
}


def _map_headers(headers):
    """Map spreadsheet headers to canonical fields using the allowlist only.

    Returns (positions, ignored_count, missing_columns). Unknown columns are counted
    as ignored and never imported.
    """
    positions = {}
    for idx, raw in enumerate(headers):
        if raw is None:
            continue
        normalized = _normalize(str(raw))
        if not normalized:
            continue
        matched = False
        for canonical, aliases in NORMALIZED_ALIASES.items():
            if canonical not in positions and normalized in aliases:
                positions[canonical] = idx
                matched = True
                break
        if not matched:
            positions.setdefault("_ignored", []).append(idx)

    ignored_count = len(positions.get("_ignored", []))
    missing_columns = [c for c in REQUIRED_COLUMNS if c not in positions]
    return positions, ignored_count, missing_columns


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _validate_row(row_number: int, data) -> ImportRow:
    name = _cell_text(data.get("name"))
    email = _cell_text(data.get("email"))
    phone = _cell_text(data.get("phone_number"))
    section = _cell_text(data.get("section"))
    semester = _cell_text(data.get("semester"))

    errors = []
    if not name:
        errors.append("Missing name")
    if not email:
        errors.append("Missing email")
    elif not EMAIL_RE.match(email):
        errors.append("Invalid email")
    if not phone:
        errors.append("Missing phone number")
    if not section:
        errors.append("Missing section")
    if not semester:
        errors.append("Missing semester")

    return ImportRow(
        row_number=row_number,
        name=name,
        email=email,
        phone_number=phone,
        section=section,
        semester=semester,
        errors=errors or None,
    )


def _parse_excel(file_bytes: bytes) -> tuple:
    if not file_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty"
        )

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read spreadsheet. Ensure it is a valid .xlsx file.",
        )

    sheet = workbook.active
    if sheet is None or sheet.max_row is None or sheet.max_row < 1:
        workbook.close()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Spreadsheet contains no data"
        )

    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        workbook.close()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Spreadsheet contains no data"
        )

    data_rows = []
    for idx, row in enumerate(rows, start=1):
        if idx > MAX_ROW_COUNT:
            workbook.close()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Spreadsheet exceeds maximum row limit of {MAX_ROW_COUNT} rows.",
            )
        data_rows.append([_cell_text(cell) for cell in row])
    workbook.close()
    return list(header), data_rows


class ImportService:
    MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_BYTES

    @classmethod
    def _read_upload(cls, file: UploadFile) -> bytes:
        filename = (file.filename or "").lower()
        if not filename.endswith(tuple(ALLOWED_EXTENSIONS)):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported file type. Please upload a .xlsx spreadsheet.",
            )

        chunks = []
        total_size = 0
        while True:
            chunk = file.file.read(8192)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > cls.MAX_FILE_SIZE_BYTES:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="File is too large. Maximum size is 10 MB.",
                )
            chunks.append(chunk)

        content = b"".join(chunks)
        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty"
            )
        return content

    @classmethod
    def _event_or_404(cls, db: Session, event_id: int) -> Event:
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Event not found")
        return event

    @classmethod
    def _row_data(cls, row, positions) -> dict:
        return {
            field: (row[positions[field]] if positions.get(field, -1) < len(row) else "")
            for field in REQUIRED_COLUMNS
        }

    @classmethod
    def preview(cls, db: Session, event_id: int, file: UploadFile) -> AttendeeImportPreview:
        cls._event_or_404(db, event_id)
        content = cls._read_upload(file)
        headers, data_rows = _parse_excel(content)

        positions, ignored_count, missing_columns = _map_headers(headers)
        if missing_columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Could not find required column"
                    f"{'s' if len(missing_columns) > 1 else ''}: "
                    f"{', '.join(DISPLAY_NAMES[c] for c in missing_columns)}"
                ),
            )

        existing_emails = {
            row.email
            for row in db.query(Attendee.email)
            .filter(Attendee.event_id == event_id)
            .all()
        }

        rows = []
        email_seen = set()
        duplicate_emails = set()
        existing_count = 0

        for row_idx, row in enumerate(data_rows, start=2):
            parsed = _validate_row(row_idx, cls._row_data(row, positions))

            if parsed.errors:
                rows.append(parsed)
                continue

            key = parsed.email.lower()
            if key in email_seen:
                duplicate_emails.add(key)
                parsed.errors = ["Duplicate email in spreadsheet"]
                rows.append(parsed)
                continue
            email_seen.add(key)
            rows.append(parsed)

        for row in rows:
            if not row.errors and row.email.lower() in existing_emails:
                row.errors = ["Email already registered for this event"]
                existing_count += 1

        valid_count = sum(1 for r in rows if not r.errors)
        invalid_count = len(rows) - valid_count

        return AttendeeImportPreview(
            filename=file.filename or "upload",
            detected_columns=list(REQUIRED_COLUMNS),
            ignored_column_count=ignored_count,
            total_rows=len(data_rows),
            valid_count=valid_count,
            invalid_count=invalid_count,
            duplicate_count=len(duplicate_emails),
            existing_count=existing_count,
            rows=rows,
            missing_columns=[],
        )

    @classmethod
    def confirmed_import(cls, db: Session, event_id: int, file: UploadFile) -> AttendeeImportResult:
        cls._event_or_404(db, event_id)
        content = cls._read_upload(file)
        headers, data_rows = _parse_excel(content)

        positions, _ignored, missing_columns = _map_headers(headers)
        if missing_columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Could not find required column"
                    f"{'s' if len(missing_columns) > 1 else ''}: "
                    f"{', '.join(DISPLAY_NAMES[c] for c in missing_columns)}"
                ),
            )

        existing_emails = {
            row.email
            for row in db.query(Attendee.email)
            .filter(Attendee.event_id == event_id)
            .all()
        }

        imported = 0
        invalid_skipped = 0
        existing_skipped = 0
        duplicate_skipped = 0
        email_seen = set()

        for row_idx, row in enumerate(data_rows, start=2):
            parsed = _validate_row(row_idx, cls._row_data(row, positions))

            if parsed.errors:
                invalid_skipped += 1
                continue

            key = parsed.email.lower()
            if key in email_seen:
                duplicate_skipped += 1
                continue
            email_seen.add(key)

            if key in existing_emails:
                existing_skipped += 1
                continue

            try:
                attendee = Attendee(
                    event_id=event_id,
                    name=parsed.name,
                    email=parsed.email,
                    phone_number=parsed.phone_number,
                    section=parsed.section,
                    semester=parsed.semester,
                )
                db.add(attendee)
                db.flush()
                db.add(
                    Attendance(
                        event_id=event_id,
                        attendee_id=attendee.id,
                        is_present=False,
                    )
                )
                imported += 1
            except Exception:
                db.rollback()
                raise

        db.commit()

        return AttendeeImportResult(
            processed=len(data_rows),
            imported=imported,
            invalid_skipped=invalid_skipped,
            existing_skipped=existing_skipped,
            duplicate_skipped=duplicate_skipped,
        )
